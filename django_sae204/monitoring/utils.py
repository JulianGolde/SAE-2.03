"""
Fonctions utilitaires (boîte à outils) :
 - parse_csv      : lecture robuste d'un fichier CSV de mesures ;
 - build_workbook : génération d'un classeur Excel (.xlsx) avec graphique.
"""
import csv              # lecture de fichiers CSV
import io               # pour traiter une chaîne comme un fichier (StringIO)
from datetime import datetime

# openpyxl = bibliothèque pour créer des fichiers Excel
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference     # graphique en courbe
from openpyxl.styles import Alignment, Font, PatternFill  # mise en forme des cellules
from openpyxl.utils import get_column_letter        # convertit 1 -> "A", 2 -> "B", etc.


# Formats de date acceptés à l'import (on essaie le FR d'abord, puis l'ISO)
_DATE_FORMATS = (
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%d/%m/%Y",
    "%Y-%m-%d",
)

# Noms de colonnes acceptés dans l'en-tête du CSV.
# clé interne -> toutes les variantes possibles (pour être tolérant).
_ALIASES = {
    "id_capteur": ("id_capteur", "id", "capteur", "idcapteur", "sensor", "sonde"),
    "piece": ("piece", "pièce", "lieu", "salle", "room"),
    "nom": ("nom", "name", "libelle", "libellé"),
    "emplacement": ("emplacement", "localisation", "location"),
    "date_mesure": ("date_mesure", "date", "datetime", "horodatage", "timestamp"),
    "heure": ("heure", "time", "hour"),
    "temperature": ("temperature", "température", "temp", "valeur", "value"),
}


def _normalise_entete(nom):
    """Nettoie un nom de colonne : enlève espaces, met en minuscules, retire le BOM."""
    return (nom or "").strip().lower().lstrip("﻿")


def _construire_mapping(entete):
    """Associe chaque clé interne (id_capteur, temperature...) au NUMÉRO de colonne du CSV."""
    cols = [_normalise_entete(c) for c in entete]   # en-tête nettoyée
    mapping = {}
    for cle, variantes in _ALIASES.items():         # pour chaque champ attendu
        for i, col in enumerate(cols):              # on cherche dans les colonnes
            if col in variantes:                    # une variante reconnue ?
                mapping[cle] = i                    # on retient l'index de colonne
                break
    return mapping


def _parse_date(valeur_date, valeur_heure=""):
    """Transforme une date (et éventuellement une heure séparée) en objet datetime, ou None."""
    txt = valeur_date.strip()
    if valeur_heure and valeur_heure.strip():       # si l'heure est dans une colonne à part
        txt = f"{txt} {valeur_heure.strip()}"       # on recolle "date heure"
    for fmt in _DATE_FORMATS:                        # on essaie chaque format connu
        try:
            return datetime.strptime(txt, fmt)      # premier format qui marche -> on renvoie
        except ValueError:
            continue                                 # sinon on essaie le suivant
    return None                                      # aucun format ne convient


def _parse_temperature(valeur):
    """Transforme un texte en nombre flottant. Gère la virgule décimale. Renvoie None si invalide."""
    try:
        return float(valeur.strip().replace(",", "."))  # "26,35" -> 26.35
    except (ValueError, AttributeError):
        return None


def parse_csv(fichier):
    """
    Lit un fichier CSV téléversé et renvoie un tuple (mesures, erreurs).

    mesures : liste de dictionnaires {id_capteur, piece, nom, emplacement, date_mesure, temperature}
    erreurs : liste de messages décrivant les lignes ignorées
    """
    contenu = fichier.read()                         # lit tout le fichier (octets ou texte)
    if isinstance(contenu, bytes):                   # si ce sont des octets, on décode en texte
        try:
            contenu = contenu.decode("utf-8-sig")    # utf-8 (en retirant un éventuel BOM)
        except UnicodeDecodeError:
            contenu = contenu.decode("latin-1")      # repli pour les fichiers Windows anciens

    # --- Détection automatique du séparateur (',' ';' ou tabulation) ---
    echantillon = contenu[:2048]                     # on regarde le début du fichier
    try:
        delimiteur = csv.Sniffer().sniff(echantillon, delimiters=",;\t").delimiter
    except csv.Error:                                # si la détection échoue
        # on devine : si plus de ';' que de ',', c'est ';' (Excel FR), sinon ','
        delimiteur = ";" if echantillon.count(";") > echantillon.count(",") else ","

    lecteur = csv.reader(io.StringIO(contenu), delimiter=delimiteur)  # lecteur CSV
    lignes = list(lecteur)                            # toutes les lignes en mémoire

    mesures, erreurs = [], []
    if not lignes:                                    # fichier vide
        return mesures, ["Fichier vide."]

    mapping = _construire_mapping(lignes[0])          # 1re ligne = en-tête -> numéros de colonnes
    # Les 3 colonnes ci-dessous sont obligatoires :
    manquants = [c for c in ("id_capteur", "date_mesure", "temperature") if c not in mapping]
    if manquants:                                     # s'il en manque, on arrête
        erreurs.append(
            "Colonnes obligatoires introuvables : " + ", ".join(manquants)
            + ". En-tête lue : " + ", ".join(lignes[0])
        )
        return mesures, erreurs

    # On parcourt les lignes de données (à partir de la 2e ; n démarre à 2 pour les messages)
    for n, ligne in enumerate(lignes[1:], start=2):
        if not any(cell.strip() for cell in ligne):  # ligne entièrement vide
            continue
        try:                                          # on lit les cellules par leur numéro
            id_capteur = ligne[mapping["id_capteur"]].strip()
            date_brute = ligne[mapping["date_mesure"]]
            heure = ligne[mapping["heure"]] if "heure" in mapping else ""
            temp = ligne[mapping["temperature"]]
        except IndexError:                            # ligne trop courte
            erreurs.append(f"Ligne {n} : nombre de colonnes incorrect, ignorée.")
            continue

        if not id_capteur:                            # id obligatoire
            erreurs.append(f"Ligne {n} : id_capteur manquant, ignorée.")
            continue

        date_mesure = _parse_date(date_brute, heure)  # conversion de la date
        if date_mesure is None:
            erreurs.append(f"Ligne {n} : date invalide ('{date_brute}'), ignorée.")
            continue

        temperature = _parse_temperature(temp)        # conversion de la température
        if temperature is None:
            erreurs.append(f"Ligne {n} : température invalide ('{temp}'), ignorée.")
            continue

        # Colonnes optionnelles (on vérifie qu'elles existent ET que la cellule est présente)
        piece = ligne[mapping["piece"]].strip() if "piece" in mapping and mapping["piece"] < len(ligne) else ""
        nom = ligne[mapping["nom"]].strip() if "nom" in mapping and mapping["nom"] < len(ligne) else ""
        emplacement = (
            ligne[mapping["emplacement"]].strip()
            if "emplacement" in mapping and mapping["emplacement"] < len(ligne)
            else ""
        )

        # Ligne valide -> on l'ajoute (avec des valeurs par défaut si optionnelles vides)
        mesures.append(
            {
                "id_capteur": id_capteur,
                "piece": piece or "Inconnue",
                "nom": nom or id_capteur,
                "emplacement": emplacement or (piece or "Inconnu"),
                "date_mesure": date_mesure,
                "temperature": temperature,
            }
        )

    return mesures, erreurs


def build_workbook(mesures, seuil_haut, seuil_bas, titre="Mesures de température"):
    """
    Construit un classeur Excel à partir d'une liste de mesures (objets Donnee).
    Retourne un objet openpyxl.Workbook (que la vue enregistre dans la réponse HTTP).
    """
    wb = Workbook()                                   # nouveau classeur (1 feuille par défaut)

    # --- Feuille 1 : Mesures --------------------------------------------------
    ws = wb.active                                    # la feuille active
    ws.title = "Mesures"                              # on la renomme

    entetes = ["Capteur (ID)", "Nom", "Pièce", "Date", "Température (°C)", "État"]
    ws.append(entetes)                               # 1re ligne = en-têtes

    # Mise en forme de la ligne d'en-tête (fond bleu, texte blanc, centré)
    entete_fill = PatternFill("solid", fgColor="1F4E78")
    entete_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(entetes) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = entete_fill
        c.font = entete_font
        c.alignment = Alignment(horizontal="center")

    rouge = PatternFill("solid", fgColor="F8CBAD")   # fond pour les mesures HAUTE
    bleu = PatternFill("solid", fgColor="BDD7EE")    # fond pour les mesures BASSE

    mesures = list(mesures)                           # on fige la liste (on va la parcourir 2 fois)
    ligne = 2                                         # on écrit à partir de la 2e ligne
    for m in mesures:
        # Détermine l'état + la couleur de fond selon les seuils
        if m.temperature >= seuil_haut:
            etat, fill = "HAUTE", rouge
        elif m.temperature <= seuil_bas:
            etat, fill = "BASSE", bleu
        else:
            etat, fill = "OK", None

        ws.cell(row=ligne, column=1, value=m.capteur_id)      # ID capteur
        ws.cell(row=ligne, column=2, value=m.capteur.nom)     # nom
        ws.cell(row=ligne, column=3, value=m.capteur.piece)   # pièce
        cdate = ws.cell(row=ligne, column=4, value=m.date_mesure)  # date
        cdate.number_format = "DD/MM/YYYY HH:MM:SS"            # format d'affichage de la date
        ws.cell(row=ligne, column=5, value=round(m.temperature, 2))  # température
        ws.cell(row=ligne, column=6, value=etat)              # état
        if fill:                                               # si critique : colorer toute la ligne
            for col in range(1, 7):
                ws.cell(row=ligne, column=col).fill = fill
        ligne += 1

    # Largeurs de colonnes (lisibilité) + figer la ligne d'en-tête
    largeurs = [16, 18, 16, 22, 16, 10]
    for i, w in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"                            # l'en-tête reste visible au défilement

    # --- Graphique en courbe (si au moins 2 points) ---------------------------
    n = len(mesures)
    if n >= 2:
        chart = LineChart()
        chart.title = titre
        chart.style = 12
        chart.y_axis.title = "Température (°C)"
        chart.x_axis.title = "Date"
        chart.height = 10
        chart.width = 26

        # Référence aux données : colonne 5 (température), de la ligne 1 (titre) à n+1
        donnees = Reference(ws, min_col=5, min_row=1, max_row=n + 1)
        # Référence aux catégories (axe X) : colonne 4 (date), lignes 2 à n+1
        cats = Reference(ws, min_col=4, min_row=2, max_row=n + 1)
        chart.add_data(donnees, titles_from_data=True)  # 1re cellule = titre de la série
        chart.set_categories(cats)
        chart.x_axis.number_format = "DD/MM HH:MM"
        chart.x_axis.majorTimeUnit = "days"
        ws.add_chart(chart, "H2")                     # place le graphique à partir de la cellule H2

    # --- Feuille 2 : Synthèse par capteur ------------------------------------
    ws2 = wb.create_sheet("Synthèse")                # nouvelle feuille
    ws2.append(["Capteur", "Pièce", "Nb mesures", "Min (°C)", "Moy (°C)", "Max (°C)", "Nb critiques"])
    for col in range(1, 8):                           # même mise en forme d'en-tête
        c = ws2.cell(row=1, column=col)
        c.fill = entete_fill
        c.font = entete_font
        c.alignment = Alignment(horizontal="center")

    # Calcul des statistiques par capteur
    stats = {}
    for m in mesures:
        # setdefault : crée l'entrée du capteur la 1re fois
        s = stats.setdefault(
            m.capteur_id,
            {"piece": m.capteur.piece, "nom": m.capteur.nom, "temps": [], "crit": 0},
        )
        s["temps"].append(m.temperature)             # on accumule les températures
        if m.temperature >= seuil_haut or m.temperature <= seuil_bas:
            s["crit"] += 1                            # on compte les critiques

    r = 2
    for cid, s in sorted(stats.items()):             # une ligne de synthèse par capteur
        temps = s["temps"]
        ws2.cell(row=r, column=1, value=s["nom"])
        ws2.cell(row=r, column=2, value=s["piece"])
        ws2.cell(row=r, column=3, value=len(temps))                       # nb mesures
        ws2.cell(row=r, column=4, value=round(min(temps), 2))            # min
        ws2.cell(row=r, column=5, value=round(sum(temps) / len(temps), 2))  # moyenne
        ws2.cell(row=r, column=6, value=round(max(temps), 2))            # max
        ws2.cell(row=r, column=7, value=s["crit"])                       # nb critiques
        r += 1
    for i, w in enumerate([18, 16, 12, 10, 10, 10, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    return wb                                         # le classeur prêt à être enregistré
